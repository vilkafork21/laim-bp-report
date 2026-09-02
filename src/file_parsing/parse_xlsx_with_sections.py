from __future__ import annotations

import json
import re
import uuid
from typing import Any, Dict, List, Optional, Union

import openpyxl


class DocumentSection:
    def __init__(self, title: str = "", level: int = 0):
        self.id: str = str(uuid.uuid4())
        self.title: str = title
        self.level: int = level
        self.content: List[Union[str, Dict[str, Any]]] = []
        self.subsections: List[DocumentSection] = []
        self.tags: List[str] = []

    def to_dict(self) -> Dict[str, Any]:
        data = {
            "id": self.id,
            "title": self.title,
            "level": self.level,
            "tags": self.tags,
            "content": self.content,
            "subsections": [sub.to_dict() for sub in self.subsections],
        }
        if self.level == 0:
            data["full_plain_text"] = self.get_all_plain_text()
        return data

    def to_json(self) -> str:
        return json.dumps(self.to_dict(), ensure_ascii=False, indent=2)

    @staticmethod
    def from_json(json_str: str) -> DocumentSection:
        def _from_dict(data: Dict[str, Any]) -> DocumentSection:
            section = DocumentSection(title=data.get("title", ""), level=data.get("level", 0))
            section.id = data.get("id", str(uuid.uuid4()))
            section.tags = data.get("tags", [])
            section.content = data.get("content", [])
            for sub_data in data.get("subsections", []):
                section.subsections.append(_from_dict(sub_data))
            return section

        data = json.loads(json_str)
        return _from_dict(data)

    def flatten(self, parent_titles: List[str] = None) -> List[Dict[str, Any]]:
        if parent_titles is None:
            parent_titles = []
        full_title = (
            " > ".join(parent_titles + [self.title]) if self.title else " > ".join(parent_titles)
        )
        flattened = [
            {
                "id": self.id,
                "full_title": full_title,
                "level": self.level,
                "tags": self.tags,
                "content": self.content,
            }
        ]
        for sub in self.subsections:
            flattened.extend(sub.flatten(parent_titles + [self.title]))
        return flattened

    def get_tree(self, indent: int = 0) -> str:
        result = "  " * indent + f"{self.title} (Level {self.level}, id: {self.id})\n"
        for sub in self.subsections:
            result += sub.get_tree(indent + 1)
        return result

    def print_tree(self, indent: int = 0):
        print(self.get_tree(indent))

    def get_full_text(self) -> str:
        texts = []
        for item in self.content:
            if isinstance(item, str):
                texts.append(item)
            elif isinstance(item, dict) and "table_text" in item:
                texts.append(item["table_text"])
        return "\n".join(texts)

    def search(self, keyword: str) -> List[Dict[str, Any]]:
        results = []
        lower_kw = keyword.lower()
        if lower_kw in self.title.lower() or lower_kw in self.get_full_text().lower():
            results.append(self.to_dict())
        for sub in self.subsections:
            results.extend(sub.search(keyword))
        return results

    def search_by_regex(self, pattern: str) -> List[Dict[str, Any]]:
        results = []
        try:
            regex = re.compile(pattern, re.IGNORECASE)
        except re.error as e:
            return results
        if regex.search(self.title) or regex.search(self.get_full_text()):
            results.append(self.to_dict())
        for sub in self.subsections:
            results.extend(sub.search_by_regex(pattern))
        return results

    def filter_by_tag(self, tag: str) -> List[Dict[str, Any]]:
        results = []
        if tag in self.tags:
            results.append(self.to_dict())
        for sub in self.subsections:
            results.extend(sub.filter_by_tag(tag))
        return results

    def find_by_id(self, search_id: str) -> Optional[DocumentSection]:
        if self.id == search_id:
            return self
        for sub in self.subsections:
            found = sub.find_by_id(search_id)
            if found:
                return found
        return None

    def search_by_title(self, title: str) -> List[Dict[str, Any]]:
        """
        Search sections by title using case-insensitive partial matching.

        Args:
            title: Search query string

        Returns:
            List of matching sections as dictionaries
        """
        results = []
        title_lower = title.lower()

        # Check current section
        if title_lower in self.title.lower():
            results.append(self.to_dict())

        # Recursively search subsections
        for sub in self.subsections:
            results.extend(sub.search_by_title(title))

        return results

    def get_all_plain_text(
        self, structured: bool = False, titles: Optional[List[str]] = None
    ) -> str:
        """
        Recursively gathers text from this section and all nested subsections.

        Args:
            structured: If True, formats sections with headers and separators in a structured format.
                       If False (default), returns simple plain text with minimal formatting.
            titles: List of section titles to include. If None, includes all sections.
                   Works for both structured and non-structured formats.

        Returns:
            str: Combined text content with formatting based on the specified parameters
        """

        if structured:
            # Structured format: iterate over worksheets or specified sections
            text = ""

            # If this is the root section, process its subsections (worksheets)
            if self.level == 0:
                # Filter sections if titles are provided, otherwise use all worksheets
                sections_to_process = []
                if titles:
                    # Find sections with matching titles
                    for title in titles:
                        matches = self.search_by_title(title)
                        if matches:
                            # Get the actual section object for each match
                            for match in matches:
                                section = self.find_by_id(match["id"])
                                if section:
                                    sections_to_process.append(section)
                else:
                    # Use all direct subsections (worksheets)
                    sections_to_process = self.subsections

                # Process each selected section
                for section in sections_to_process:
                    text += f"\n\n### {section.title} ###\n\n"
                    text += str(section.flatten())
                    text += f"\n\n{'-' * 50}\n\n"
            else:
                # If this is not the root, just add this section's content
                if self.title:
                    text += f"\n\n### {self.title} ###\n\n"
                    text += str(self.flatten())
                    text += f"\n\n{'-' * 50}\n\n"

            return text.strip()

        else:
            # Simple plain text format with title filtering
            if titles and self.level == 0:
                # Only used at root level when titles specified
                sections_text = []

                # Find and process sections with matching titles
                for title in titles:
                    matches = self.search_by_title(title)
                    for match in matches:
                        section = self.find_by_id(match["id"])
                        if section:
                            if section.title:
                                section_content = f"# {section.title}\n{section.get_full_text()}"
                                sections_text.append(section_content)
                            else:
                                sections_text.append(section.get_full_text())

                return "\n\n".join([txt for txt in sections_text if txt]).strip()
            else:
                # Regular recursive behavior for non-root or when no titles specified
                sections_text = []

                # Add current section title and content
                if self.title:
                    section_content = f"# {self.title}\n{self.get_full_text()}"
                    sections_text.append(section_content)
                else:
                    full_text = self.get_full_text()
                    if full_text:
                        sections_text.append(full_text)

                # Add subsections recursively
                for sub in self.subsections:
                    subsection_text = sub.get_all_plain_text(structured=False)
                    if subsection_text:
                        sections_text.append(subsection_text)

                return "\n\n".join([txt for txt in sections_text if txt]).strip()


def parse_sheet_as_table(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
) -> Optional[Dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return None

    headers = [str(cell) if cell is not None else f"Column {i+1}" for i, cell in enumerate(rows[0])]

    table_dict = []
    table_text_rows = []
    for row in rows[1:]:
        row_values = [str(cell) if cell is not None else "" for cell in row]
        table_text_rows.append("\t".join(row_values).strip())

        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        table_dict.append(row_dict)

    table_text = "\n".join(table_text_rows)
    return {"table_text": table_text, "table_dict": table_dict}


def parse_xlsx_to_sections(file_path: str) -> DocumentSection:
    root_section = DocumentSection(title="", level=0)

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        raise

    for sheet in workbook.worksheets:
        section = DocumentSection(title=sheet.title, level=1)
        section.tags.append("worksheet")

        table = parse_sheet_as_table(sheet)
        if table:
            section.content.append(table)
        else:
            for row in sheet.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                row_text = "\t".join(row_values).strip()
                if row_text:
                    section.content.append(row_text)

        root_section.subsections.append(section)

    return root_section
